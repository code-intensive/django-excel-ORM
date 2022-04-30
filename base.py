import openpyxl
from .spreadsheets_settings import WORKBOOK_BASE_DIR
from typing import Generator
from .exceptions import WorkSheetNotFoundError
from django.db.models import Model


class DjangoExcelToDB:
    '''
    Class for populating database using django's ORM
    '''
    DEFAULT_WORKBOOK_PATH: str = WORKBOOK_BASE_DIR.joinpath(
        'stocks_cleaned.xlsx').resolve().__str__()
    # fields_to_extract_from_db = []

    model: Model = None

    def __init__(self, *, workbook_path: str | None = None, cleaned_sheet_name: str = 'cleaned sheet'):
        WORKBOOK_PATH = workbook_path or self.DEFAULT_WORKBOOK_PATH
        self.__workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        self.__run_checks__(cleaned_sheet_name)

    def __run_checks__(self, cleaned_sheet_name) -> None:
        """Run core checks on to ensure:

        * `fields_to_extract_from_db` (`cls property`) is not empty
        * `Meta` class's `model` property is specified

        Args:
            `cleaned_sheet_name` (`str`): Sheet to be used for 
            population of database

        Raises:
            ExcelToDBException.WorkSheetNotFoundError: Raised when the 
            specified worksheet is not found in the loaded workbook
        """
        # Check if the provided cleaned_sheet_name argument is present in
        # the list of existing workbook sheets if not raise exception
        if not cleaned_sheet_name in self.__workbook.sheetnames:
            raise WorkSheetNotFoundError(
                'Worksheet named "%s" does not exist, '
                'Could you have made a typo?' % cleaned_sheet_name.upper()
            )
        else:
            self.__worksheet = self.__workbook[cleaned_sheet_name]

        # Ensure fields are not empty
        # assert len(
        #     self.fields_to_extract_from_db) > 0, (
        #         '"fields_to_extract_from_db" is missing, kindly supply field(s)'
        # )

        # Make sure a django model subclass is passed in as model of Meta class
        assert self.model is not None and issubclass(self.model, Model), (
            'Model attribute missing or invalid model class passed, '
            'Kindly define a model attribute - '
            '*(This should be a django models.Model subclass)'
        )

    def migrate_to_db(self):
        """Extracts values and pushes them to the database
        """
        _objects = list(self._extract_values())
        self._push_to_db(
            _objects
        )
        success_msg = '%d object(s) of %s successfully created and pushed to the database' % (
            len(_objects), self.model.__class__
        )
        print(success_msg)

    def _extract_values(self) -> Generator:
        # TODO use a header attribute of array to dynamically get the sheet's
        # start point so as to know where the main extraction begins
        if self.__worksheet.cell(1, 1).value.__str__().isalpha():
            # Delete header upon finding it
            self.__worksheet.delete_rows(0)
        # Yield each row's unpacked value for processing
        for row in self.__worksheet.values:
            yield self.model(*row)

    def _push_to_db(self, _objects):
        """Populates database with objects created from the
        specified excel sheet
        """
        self.model.objects.bulk_create(_objects)
